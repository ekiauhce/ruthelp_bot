from io import BytesIO
from openpyxl import (
    Workbook,
    load_workbook
)
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from database.db import (
    get_applications_field_names,
    get_categories_list,
    get_applications,
    set_application_ok
)

style_red = DifferentialStyle(
    fill=PatternFill(bgColor="FFCCCB")
)

style_green = DifferentialStyle(
    fill=PatternFill(bgColor="90EE90")   
)

#В случае изменения структуры дб -
#изменить индекс
rule_red = Rule(
    type="expression",
    dxf=style_red,
    formula=["$J1=0"]
)

rule_green = Rule(
    type="expression",
    dxf=style_green,
    formula=["$J1=1"]
)

def init_wb() -> Workbook:
    wb = Workbook()
    wb.active.title = "1"
    wb.create_sheet("2")
    wb.create_sheet("3")
    wb.create_sheet("4")
    for ws in wb:
        ws.append(get_applications_field_names())
    return wb


def download_from_db() -> BytesIO:
    """
    Возвращает байтовый буфер с таблицой бд
    в формате xlsx
    """
    wb = init_wb()
    categories = get_categories_list()
    for app in get_applications():
        categories = get_categories_list()
        if app[1] in categories[:3]:
            wb["1"].append(app)
        elif app[1] in categories[3:6] + categories[8:10]:
            wb["2"].append(app)
        elif app[1] in categories[6:8]:
            wb["3"].append(app)
        elif app[1] == categories[10]:
            wb["4"].append(app)
    
    for ws in wb:
        ws.conditional_formatting.add(
            ws.dimensions,
            rule_red
        )
        ws.conditional_formatting.add(
            ws.dimensions,
            rule_green
        )
        #В случае изменения структуры дб -
        #изменить индекс
        ws.auto_filter.ref = f"J1:J{ws.max_row}"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def upload_to_db(buffer: BytesIO):
    """
    Записывает изменения, сделанные в xlsx в файле,
    в базу данных
    """
    wb = load_workbook(filename=buffer)
    for ws in wb:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[9].value == 1:
                set_application_ok(row[0].value)
